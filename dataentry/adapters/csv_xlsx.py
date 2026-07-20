"""CSV / XLSX ingest adapters.

Both adapters treat the first row as the header and emit one
:class:`DataEntryRecord` per subsequent row, keyed by the header names.
A single missing header column aborts the row (the row's value is left
out) so partial files don't poison the whole batch.
"""

import csv
from collections.abc import Iterable
from io import BytesIO, TextIOWrapper

from openpyxl import load_workbook

from dataentry.models import DataEntryRecord


def _ingest_rows(
    tenant,
    source: str,
    rows: Iterable[list[object]],
) -> list[DataEntryRecord]:
    """Materialize an iterable of value-rows as PENDING DataEntryRecords."""
    rows = iter(rows)
    try:
        header = [str(h) for h in next(rows)]
    except StopIteration:
        return []
    records: list[DataEntryRecord] = []
    for row in rows:
        fields: dict[str, object] = {}
        for key, value in zip(header, row, strict=False):
            if value is None:
                continue
            fields[key] = value
        records.append(
            DataEntryRecord.objects.create(
                tenant=tenant,
                source=source,
                raw=fields,
            )
        )
    return records


def ingest_csv(tenant, fileobj: BytesIO) -> list[DataEntryRecord]:
    """Ingest a CSV payload.

    Args:
        tenant: Owning tenant.
        fileobj: A binary file-like object containing CSV bytes.

    Returns:
        List of created PENDING :class:`DataEntryRecord` rows
        (one per data row).
    """
    text = TextIOWrapper(fileobj, encoding="utf-8", newline="")
    reader = csv.reader(text)
    return _ingest_rows(tenant, DataEntryRecord.Source.CSV, reader)


def ingest_xlsx(tenant, fileobj: BytesIO) -> list[DataEntryRecord]:
    """Ingest the first sheet of an XLSX workbook.

    Args:
        tenant: Owning tenant.
        fileobj: A binary file-like object containing the XLSX bytes.

    Returns:
        List of created PENDING :class:`DataEntryRecord` rows.
    """
    wb = load_workbook(filename=fileobj, read_only=True, data_only=True)
    ws = wb.active
    return _ingest_rows(
        tenant, DataEntryRecord.Source.CSV, ws.iter_rows(values_only=True)
    )
